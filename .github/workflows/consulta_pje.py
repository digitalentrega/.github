import requests
import pandas as pd
from datetime import datetime
import os
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("consulta_pje.log"),
        logging.StreamHandler()
    ]
)

# Diretório para salvar os arquivos Excel
DIRETORIO_SAIDA = "consulta_pje"
os.makedirs(DIRETORIO_SAIDA, exist_ok=True)

# Configurações da consulta
NUMERO_OAB = "46470"
UF_OAB = "PR"
ITENS_POR_PAGINA = 100

def formatar_processo(numero):
    if "-" not in numero and "." not in numero:
        try:
            return f"{numero[:7]}-{numero[7:9]}.{numero[9:13]}.{numero[13:14]}.{numero[14:16]}.{numero[16:]}"
        except:
            return numero
    return numero

def consultar_api(data_inicio, data_fim):
    url = "https://comunicaapi.pje.jus.br/api/v1/comunicacao"
    
    params = {
        "numeroOab": NUMERO_OAB,
        "ufOab": UF_OAB,
        "dataDisponibilizacaoInicio": data_inicio,
        "dataDisponibilizacaoFim": data_fim,
        "itensPorPagina": ITENS_POR_PAGINA
    }
    
    headers = {
        "accept": "application/json"
    }
    
    try:
        logging.info(f"Consultando API para o período de {data_inicio} a {data_fim}")
        response = requests.get(url, params=params, headers=headers)
        response.raise_for_status()
        return response.json()
    except requests.exceptions.RequestException as e:
        logging.error(f"Erro ao consultar API: {e}")
        return None

def processar_dados(dados):
    if not dados or "items" not in dados:
        logging.warning("Nenhum dado encontrado ou formato inválido")
        return pd.DataFrame()

    registros = []
    for item in dados["items"]:
        destinatarios = ", ".join([d["nome"] for d in item.get("destinatarios", [])])
        advogados = []
        for adv in item.get("destinatarioadvogados", []):
            if "advogado" in adv and adv["advogado"]:
                advogado_info = adv["advogado"]
                advogados.append(f"{advogado_info.get('nome', '')} (OAB {advogado_info.get('numero_oab', '')}/{advogado_info.get('uf_oab', '')})")
        advogados_str = ", ".join(advogados)
        numero_processo_formatado = item.get("numeroprocessocommascara", item.get("numero_processo", ""))
        if not numero_processo_formatado:
            numero_processo_formatado = formatar_processo(item.get("numero_processo", ""))

        registro = {
            "ID": item.get("id"),
            "Data Disponibilização": item.get("data_disponibilizacao") or item.get("datadisponibilizacao"),
            "Tribunal": item.get("siglaTribunal"),
            "Tipo Comunicação": item.get("tipoComunicacao"),
            "Órgão": item.get("nomeOrgao"),
            "Número do Processo": numero_processo_formatado,
            "Classe": item.get("nomeClasse"),
            "Texto": item.get("texto"),
            "Meio": item.get("meiocompleto", item.get("meio")),
            "Link": item.get("link"),
            "Status": item.get("status"),
            "Destinatários": destinatarios,
            "Advogados": advogados_str,
        }
        registros.append(registro)

    return pd.DataFrame(registros)

def exportar_para_excel(df, data):
    if df.empty:
        logging.warning("Nenhum dado para exportar")
        return

    nome_arquivo = f"consulta_pje_{data}.xlsx"
    caminho_arquivo = os.path.join(DIRETORIO_SAIDA, nome_arquivo)

    try:
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Consulta PJe')
            workbook = writer.book
            worksheet = writer.sheets['Consulta PJe']
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).apply(len).max(), len(col)) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 50)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            worksheet.auto_filter.ref = worksheet.dimensions
        logging.info(f"Dados exportados com sucesso para {caminho_arquivo}")
        return caminho_arquivo
    except Exception as e:
        logging.error(f"Erro ao exportar para Excel: {e}")
        return None

def main():
    data = datetime.now().strftime("%Y-%m-%d")
    resultado = consultar_api(data, data)
    if not resultado:
        logging.error("Falha na consulta à API")
        return
    if resultado.get("status") != "success":
        logging.error(f"Consulta retornou erro: {resultado.get('message')}")
        return
    logging.info(f"Encontrados {resultado.get('count', 0)} registros para a data {data}")
    df = processar_dados(resultado)
    if not df.empty:
        caminho_arquivo = exportar_para_excel(df, data)
        if caminho_arquivo:
            print(f"Consulta concluída com sucesso. Arquivo salvo em: {caminho_arquivo}")
    else:
        logging.warning("Nenhum dado para exportar após processamento")

if __name__ == "__main__":
    main()
